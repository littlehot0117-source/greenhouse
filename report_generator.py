import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import database

def generate_monthly_report(year, month):
    """
    產生月報表 Excel 檔案，並儲存在 reports 專案目錄下。
    返回產生的檔案絕對路徑。
    """
    # 取得資料
    data = database.get_monthly_report_data(year, month)
    
    # 建立 DataFrame
    if not data:
        # 如果沒有資料，建立一個空的 DataFrame 結構
        df_all = pd.DataFrame(columns=[
            "溫室名稱", "品項名稱", "品項編碼", "單位", 
            "期初庫存", "本月進庫", "本月出庫", "期末庫存"
        ])
    else:
        df_all = pd.DataFrame(data)
        # 重新命名與挑選欄位
        df_all = df_all.rename(columns={
            "greenhouse_name": "溫室名稱",
            "item_name": "品項名稱",
            "item_sku": "品項編碼",
            "item_unit": "單位",
            "beginning_stock": "期初庫存",
            "month_in": "本月進庫",
            "month_out": "本月出庫",
            "ending_stock": "期末庫存"
        })
        # 挑選要呈現的欄位
        df_all = df_all[[
            "溫室名稱", "品項名稱", "品項編碼", "單位", 
            "期初庫存", "本月進庫", "本月出庫", "期末庫存"
        ]]
        
    # 確保 reports 目錄存在
    project_dir = os.path.dirname(os.path.abspath(__file__))
    reports_dir = os.path.join(project_dir, "reports")
    os.makedirs(reports_dir, exist_ok=True)
    
    file_name = f"溫室庫存月報表_{year}_{month}.xlsx"
    file_path = os.path.join(reports_dir, file_name)
    
    # 使用 ExcelWriter 寫入多個 Sheet
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        # 1. 寫入總表
        df_all.to_excel(writer, sheet_name="庫存匯總表", index=False)
        
        # 2. 依溫室分頁寫入
        greenhouses = database.get_greenhouses()
        for gh in greenhouses:
            gh_name = gh["name"]
            if not df_all.empty:
                df_gh = df_all[df_all["溫室名稱"] == gh_name].copy()
                if not df_gh.empty:
                    # 溫室分頁不需要再重複顯示「溫室名稱」欄位
                    df_gh = df_gh.drop(columns=["溫室名稱"])
                else:
                    df_gh = pd.DataFrame(columns=["品項名稱", "品項編碼", "單位", "期初庫存", "本月進庫", "本月出庫", "期末庫存"])
            else:
                df_gh = pd.DataFrame(columns=["品項名稱", "品項編碼", "單位", "期初庫存", "本月進庫", "本月出庫", "期末庫存"])
                
            df_gh.to_excel(writer, sheet_name=gh_name, index=False)
            
        # 取得 openpyxl workbook 以便進行樣式美化
        workbook = writer.book
        
        # 定義樣式
        header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="微軟正黑體", size=10)
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid") # 深綠色
        zebra_fill = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid") # 淺綠色條紋
        
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        # 格式化每一個工作表
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # 設定行高
            worksheet.row_dimensions[1].height = 28
            for r in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[r].height = 20
                
            # 套用表頭樣式
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = thin_border
                
            # 套用資料樣式
            for row in range(2, worksheet.max_row + 1):
                is_zebra = (row % 2 == 0)
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    if is_zebra:
                        cell.fill = zebra_fill
                        
                    # 依據資料型態設定對齊與格式
                    header_name = worksheet.cell(row=1, column=col).value
                    val = cell.value
                    
                    # 數值欄位靠右對齊，並設定千分位
                    if header_name in ["期初庫存", "本月進庫", "本月出庫", "期末庫存"]:
                        cell.alignment = align_right
                        if isinstance(val, (int, float)):
                            cell.number_format = '#,##0.00' if isinstance(val, float) else '#,##0'
                    elif header_name in ["單位", "品項編碼"]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
                        
            # 自動調整欄寬
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    # 處理中文字元長度 (中文計為 2 個字元長度)
                    val_len = sum(2 if ord(char) > 127 else 1 for char in val_str)
                    if val_len > max_len:
                        max_len = val_len
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            # 確保啟用網格線
            worksheet.views.sheetView[0].showGridLines = True
            
    return file_path

if __name__ == "__main__":
    # 測試生成月報表
    database.init_db()
    # 插入一些測試資料以確保測試正常
    items = database.get_items()
    if not items:
        database.add_item("測試有機肥", "F-001", "包", "測試用")
        database.add_item("測試種子", "S-002", "公克", "測試用")
        
    g_list = database.get_greenhouses()
    i_list = database.get_items()
    
    # 新增一些測試進出庫明細
    database.add_transaction(g_list[0]["id"], i_list[0]["id"], "IN", 150, "管理員", "測試進庫", "2026-08-01 10:00:00")
    database.add_transaction(g_list[0]["id"], i_list[0]["id"], "OUT", 30, "管理員", "測試出庫", "2026-08-15 14:00:00")
    database.add_transaction(g_list[1]["id"], i_list[1]["id"], "IN", 500, "管理員", "種子入庫", "2026-08-10 09:30:00")
    
    path = generate_monthly_report("2026", "08")
    print("Report generated successfully at:", path)
